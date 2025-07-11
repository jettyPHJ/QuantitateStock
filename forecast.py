import torch
import pandas as pd
import numpy as np
from MambaStock import MambaModel
import data_process.data_set as data_set
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

file_path = '测试数据.xlsx'
model_path = 'best_mamba_model.pth'

# 设置设备
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# 获取特征维度
company_names, feature_columns = data_set.get_excel_meta(file_path,exclude_columns = data_set.exclude_columns)

# 验证测试数据和训练数据的feature_columns是否相同
if feature_columns != data_set.feature_columns:
    raise ValueError("测试数据和训练数据的特征列不匹配。")

input_dim = len(feature_columns)
# 创建并加载模型
model = MambaModel()
model.load_state_dict(torch.load(model_path, weights_only=True, map_location=device))
model.to(device)
model.eval()

# 读取数据
data_map = data_set.load_excel(file_path, company_names, feature_columns)

# 输出结果字典
df_results = {}
min_size = data_set.min_sample_size - 1

# 遍历每个公司进行预测
for company in company_names:
    company_data = data_map[company]
    num_rows = len(next(iter(company_data.values())))
    if num_rows < min_size:
        print(company, ":样本不足, 无法预测")
        continue

    preds = [None] * min_size
    for i in range(min_size, num_rows):
        sub_data = {key: val[:i+1] for key, val in company_data.items()}

        # 构造输入序列
        raw_seq = []
        norm_seq = []
        for col in feature_columns:
            arr = np.array(sub_data[col], dtype=float)
            raw_seq.append(arr)
            norm_arr = data_set.sigmoid_normalize(arr)
            norm_seq.append(norm_arr)

        origin_tensor = torch.tensor(np.stack(raw_seq, axis=1), dtype=torch.float32).to(device)
        feature_tensor = torch.tensor(np.stack(norm_seq, axis=1), dtype=torch.float32).to(device)

        origin_tensor = origin_tensor.unsqueeze(0)  # (1, seq_len, dim)
        feature_tensor = feature_tensor.unsqueeze(0)
        lengths = torch.tensor([feature_tensor.shape[1]], dtype=torch.long).to(device)

        with torch.no_grad():
            pred = model([origin_tensor[0]], feature_tensor, lengths)
            preds.append(pred.item())

    # 保存预测结果
    df_company = pd.DataFrame(company_data)
    df_company['模型预测结果'] = preds
    df_results[company] = df_company

# 直接写回原始文件，添加一列“模型预测结果”到每个 sheet
wb = load_workbook(file_path)
for company, df in df_results.items():
    if company not in wb.sheetnames:
        print(f"{company} 不在 Excel 文件中，跳过")
        continue
    ws = wb[company]
    preds = df['模型预测结果'].tolist()
    # 找到写入列的位置（紧跟最后一列）
    max_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=max_col)
    header_cell.value = '模型预测结果'
    # 设置列宽
    ws.column_dimensions[get_column_letter(max_col)].width = 9
    # 设置淡绿色填充色
    fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")
    header_cell.fill = fill
    # 设置对齐方式和自动换行
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_cell.alignment = alignment
    # 设置框线
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    header_cell.border = thin_border
    for i, pred in enumerate(preds):
        cell = ws.cell(row=i+2, column=max_col)  # Excel行号从2开始（跳过标题）
        if pred is not None:
            cell.value = pred
            # 设置单元格格式为数值，显示两位小数
            cell.number_format = '0.00'
            cell.alignment = alignment  # 设置对齐方式
            cell.border = thin_border  # 设置框线
# 保存工作簿
wb.save(file_path)
print("预测结果列已追加")